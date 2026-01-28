# jobs/views.py
from rest_framework import status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from .models import JobDescription
from .serializers import (
    JobDescriptionSerializer, 
    JobDescriptionCreateSerializer, 
    JobDescriptionUpdateSerializer
)
from .utils import process_job_description_async
import io
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class JobDescriptionListCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get(self, request):
        # Filter active jobs by default
        jobs = JobDescription.objects.filter(is_active=True)
        
        # Allow filtering by priority, company, etc.
        priority = request.query_params.get('priority')
        company = request.query_params.get('company')
        
        if priority:
            jobs = jobs.filter(priority=priority)
        if company:
            jobs = jobs.filter(company_name__icontains=company)
        
        serializer = JobDescriptionSerializer(jobs, many=True, context={'request': request})
        return Response(serializer.data)
    
    def post(self, request):
        # Only placement team and admin can create jobs
        if request.user.role == 'student':
            return Response(
                {'error': 'Permission denied'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = JobDescriptionCreateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            job = serializer.save()
            # Trigger background processing
            try:
                process_job_description_async(job.id)
            except Exception as e:
                print(f"Error processing job description: {e}")
            
            response_serializer = JobDescriptionSerializer(job, context={'request': request})
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class JobDescriptionDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        job = get_object_or_404(JobDescription, pk=pk)
        serializer = JobDescriptionSerializer(job, context={'request': request})
        return Response(serializer.data)
    
    def patch(self, request, pk):
        # Only placement team and admin can update jobs
        if request.user.role == 'student':
            return Response(
                {'error': 'Permission denied'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        job = get_object_or_404(JobDescription, pk=pk)
        serializer = JobDescriptionUpdateSerializer(job, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            response_serializer = JobDescriptionSerializer(job, context={'request': request})
            return Response(response_serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        # Only placement team and admin can delete jobs
        if request.user.role == 'student':
            return Response(
                {'error': 'Permission denied'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        job = get_object_or_404(JobDescription, pk=pk)
        job.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def job_stats(request):
    """Get job statistics for dashboard"""
    active_jobs = JobDescription.objects.filter(is_active=True).count()
    high_priority_jobs = JobDescription.objects.filter(
        is_active=True, 
        priority='high'
    ).count()
    
    return Response({
        'active_jobs': active_jobs,
        'high_priority_jobs': high_priority_jobs,
        'total_jobs': JobDescription.objects.count()
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def job_matched_candidates(request, pk):
    """Get matched candidates for a specific job, sorted by score"""
    # Only placement team and admin can view matched candidates
    if request.user.role == 'student':
        return Response(
            {'error': 'Permission denied'}, 
            status=status.HTTP_403_FORBIDDEN
        )
    
    job = get_object_or_404(JobDescription, pk=pk)
    
    from evaluations.models import Evaluation
    from evaluations.serializer import EvaluationSerializer
    
    # Get all evaluations for this job, sorted by score descending
    evaluations = Evaluation.objects.filter(
        job_description=job
    ).select_related('resume', 'resume__user').order_by('-overall_score')
    
    serializer = EvaluationSerializer(evaluations, many=True, context={'request': request})
    
    return Response({
        'job': {
            'id': job.id,
            'title': job.title,
            'company_name': job.company_name,
            'positions_required': job.positions_required
        },
        'total_candidates': evaluations.count(),
        'candidates': serializer.data
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_job_candidates_excel(request, pk):
    """Export matched candidates for a job as Excel file"""
    if request.user.role == 'student':
        return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
    
    if not EXCEL_AVAILABLE:
        return Response({'error': 'Excel export not available. Install openpyxl.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    job = get_object_or_404(JobDescription, pk=pk)
    
    from evaluations.models import Evaluation
    
    # Get export type (matched, all, or shortlist)
    export_type = request.query_params.get('type', 'matched')  # 'matched', 'all', or 'shortlist'
    min_score = int(request.query_params.get('min_score', 0))
    limit = request.query_params.get('limit')  # For first round shortlisting
    round_name = request.query_params.get('round', 'First Round')  # Round name for the report
    
    evaluations = Evaluation.objects.filter(
        job_description=job
    ).select_related('resume', 'resume__user').order_by('-overall_score')
    
    if export_type == 'matched' and min_score > 0:
        evaluations = evaluations.filter(overall_score__gte=min_score)
    elif export_type == 'shortlist' and min_score > 0:
        evaluations = evaluations.filter(overall_score__gte=min_score)
    
    # Apply limit for shortlisting
    if limit:
        evaluations = evaluations[:int(limit)]
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Set sheet title based on export type
    if export_type == 'shortlist':
        ws.title = f"{round_name} Shortlist"
    else:
        ws.title = "Candidates"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    
    # Add title row for shortlist
    start_row = 1
    if export_type == 'shortlist':
        ws.cell(row=1, column=1, value=f"{round_name} Shortlisting Report")
        ws.cell(row=1, column=1).font = title_font
        ws.merge_cells('A1:L1')
        
        ws.cell(row=2, column=1, value=f"Job: {job.title} | Company: {job.company_name}")
        ws.merge_cells('A2:L2')
        
        ws.cell(row=3, column=1, value=f"Total Shortlisted: {len(list(evaluations))} candidates")
        ws.merge_cells('A3:L3')
        
        start_row = 5
    
    # Headers
    headers = ['Rank', 'Student Name', 'Email', 'Resume File', 'Overall Score', 
               'Hard Skills', 'Soft Skills', 'Experience', 'Education', 
               'Recommendation', 'Matched Skills', 'Missing Skills']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for idx, evaluation in enumerate(evaluations, 1):
        user = evaluation.resume.user
        row = start_row + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=user.get_full_name() or user.username)
        ws.cell(row=row, column=3, value=user.email)
        ws.cell(row=row, column=4, value=evaluation.resume.file_name)
        ws.cell(row=row, column=5, value=evaluation.overall_score)
        ws.cell(row=row, column=6, value=evaluation.hard_skills_score)
        ws.cell(row=row, column=7, value=evaluation.soft_skills_score)
        ws.cell(row=row, column=8, value=evaluation.experience_score)
        ws.cell(row=row, column=9, value=evaluation.education_score)
        ws.cell(row=row, column=10, value=evaluation.recommendation.replace('_', ' ').title())
        ws.cell(row=row, column=11, value=', '.join(evaluation.matched_skills[:10]) if evaluation.matched_skills else '')
        ws.cell(row=row, column=12, value=', '.join(evaluation.missing_skills[:10]) if evaluation.missing_skills else '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    if export_type == 'shortlist':
        filename = f"{job.title.replace(' ', '_')}_{round_name.replace(' ', '_')}_shortlist.xlsx"
    else:
        filename = f"{job.title.replace(' ', '_')}_{job.company_name.replace(' ', '_')}_candidates.xlsx"
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_all_applied_resumes(request, pk):
    """Get all resumes that have been evaluated for a specific job"""
    if request.user.role == 'student':
        return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
    
    job = get_object_or_404(JobDescription, pk=pk)
    
    from evaluations.models import Evaluation
    from resumes.serializer import ResumeSerializer
    from resumes.models import Resume
    
    # Get all resumes that have evaluations for this job
    evaluated_resume_ids = Evaluation.objects.filter(
        job_description=job
    ).values_list('resume_id', flat=True)
    
    resumes = Resume.objects.filter(id__in=evaluated_resume_ids)
    serializer = ResumeSerializer(resumes, many=True, context={'request': request})
    
    return Response({
        'job': {
            'id': job.id,
            'title': job.title,
            'company_name': job.company_name
        },
        'total_applied': resumes.count(),
        'resumes': serializer.data
    })
